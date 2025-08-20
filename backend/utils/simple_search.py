import os

def load_filters(filter_file="filters.txt"):
    """Φορτώνει λέξεις-κλειδιά από το filters.txt, μία ανά γραμμή."""
    if not os.path.exists(filter_file):
        return []
    with open(filter_file, "r", encoding="utf-8") as f:
        filters = [line.strip() for line in f if line.strip()]
    return filters


def search_pdf(file_path, filters):
    from utils.file_reader import read_pdf

    text = read_pdf(file_path)
    results = {}

    for query in filters:
        idx = 0
        matches = []
        while True:
            idx = text.find(query, idx)
            if idx == -1:
                break
            matches.append(query)  # μόνο το query
            idx += len(query)
        if matches:
            results[query] = matches

    return results



def search_excel(file_path, filters):
    """Αναζήτηση Excel με βάση λίστα λέξεων-κλειδιών."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    results = {}

    for query in filters:
        matches = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell and query in str(cell):
                        matches.append({
                            "sheet": sheet,
                            "row": row_idx,
                            "col": col_idx,
                            "value": str(cell)
                        })
        if matches:
            results[query] = matches

    return results


def simple_search(file_path, filter_file="filters.txt"):
    """Κεντρική συνάρτηση που διαβάζει φίλτρα και ψάχνει PDF/Excel."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_path} not found")

    filters = load_filters(filter_file)
    if not filters:
        return {}  # Αν δεν υπάρχουν φίλτρα

    ext = file_path.lower().split('.')[-1]
    if ext == "pdf":
        return search_pdf(file_path, filters)
    elif ext in ["xls", "xlsx"]:
        return search_excel(file_path, filters)
    else:
        raise ValueError("Unsupported file type")
